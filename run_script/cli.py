import click
import csv
import json
import os
import random
import re
import numpy as np
from datetime import timedelta, date
from openpyxl import load_workbook

import settings
from object.Bank import Bank
from object.Branch import Branch, ATM
from object.Product import Product
from object.User import User

@click.group()
def cli():
    pass

@cli.command(help="Generate Main Json file")
def generate_main_file():
    user_list = User.generator_for_file(settings.OPTIONS_PATH)
    bank_number = settings.BANK_NUMBER

    branch_list = []
    branch_number = settings.BRANCH_NUMBER
    atm_list = []
    atm_number = settings.ATM_NUMBER
    product_list = []
    product_number = settings.PRODUCT_NUMBER
    bank_list = Bank.generate_from_file(bank_number)
    for bank in bank_list:
        branch_list.extend(Branch.generate_from_file(bank, branch_number))
        atm_list.extend(ATM.generate_from_file(bank, atm_number))
        product_list.extend(Product.generate_from_file(bank, product_number))

    account_list = []
    for user in user_list:
        branch_tmp = random.choice(branch_list)
        account1 = user.create_account(branch_tmp, "CURRENT", user.country, user.current)
        account1.set_behavior(income=user.current, spending_frequency =
            { "food":          user.food,
              "utility":       user.utility,
              "clothing":      user.clothing,
              "auto":          user.auto,
              "health":        user.health,
              "entertainment": user.entertainment,
              "gift":         user.gift,
              "education":     user.education,
              "fee":           user.fee },
            housing_type={
                "RENT": -user.rent
            } if user.rent > 0 else None
        )
        account_list.append(account1)

        branch_tmp = random.choice(branch_list)
        account2 = user.create_account(branch_tmp, "SAVING", user.country, user.savings)
        account2.set_behavior(income=user.savings)
        account_list.append(account2)

    months = 36
    date_start = date.today() - timedelta(days=months * 30)
    transaction_list = []
    for i in range(months):
        date_start = date_start + timedelta(days=30)
        for account in account_list:
            transaction_list.extend(account.generateTransaction(date_start))

    with open("transaction.csv", 'w') as csv_file:
        fnames = ['id', 'bank', 'counterparty', 'posted', 'new_balance', 'value', 'type', 'user', 'account_type']
        # wr = csv.DictWriter(csv_file, fieldnames=fnames, delimiter=',', quotechar='|', quoting=csv.QUOTE_MINIMAL)
        # wr.writeheader()
        wr = csv.writer(csv_file, delimiter=',')
        wr.writerow(fnames)
        for transaction in transaction_list:
            wr.writerow([
                transaction.id,
                transaction.this_account.branch.bank.id,
                transaction.other_counterparty,
                str(transaction.posted),
                transaction.new_balance,
                '{:.2f}'.format(transaction.value),
                transaction.type,
                transaction.this_account.user[0].username,
                transaction.this_account.type
            ])


    output_dir = settings.OUTPUT_PATH
    try:
        os.makedirs(output_dir)
    except:
        pass
    with open('{}sandbox_pretty.json'.format(output_dir), 'w') as outfile:
        json.dump({
            "users": user_list,
            "banks": bank_list,
            "branches": branch_list,
            "accounts": account_list,
            "atms":atm_list,
            # "counterparties":counterparty_list,
            "products":product_list,
            "transactions": transaction_list
        }, outfile, default=lambda x: x.dict(), indent=4)

    customer_list = []
    for user in user_list:
        for bank in bank_list:
            customer_list.append(user.create_customer(bank))

    with open('{}customers_pretty.json'.format(output_dir), 'w') as outfile:
        json.dump(customer_list, outfile, default=lambda x: x.dict(), indent=4)


@cli.command(help="Generate Counterparty Json file")
@click.option('--input_file', default="./input_file/OBP_counterparties_TESOBE_Hong_Kong.xlsx", required=True, help='Filename of first name')
@click.option('--city', default="Hong Kong", required=True, help='Filename of first name')
def generate_counterparty_file(input_file, city):
    wb = load_workbook(input_file)

    sheet_selected = [sheetname for sheetname in wb.sheetnames if city in sheetname ]

    import pandas as pd

    dataframe = pd.read_excel(input_file, sheet_name = sheet_selected, header=None, index_col=None, skiprows=3)

    df = pd.DataFrame()
    for i, j in dataframe.items():
        df = pd.DataFrame(j)

    df.columns= ['type', 'name', 'reference', 'value', 'frequency', 'logo','homepage']

    df.type = df["type"].apply(lambda x: re.split('_|/',x)[0].lower())
    #df.frequency = df["frequency"].apply(lambda x: str("0") if x is np.nan else frequecy_mapping[x.lower()])

    df = pd.DataFrame({
        "type":df["type"].apply(lambda x: re.split('_|/',x)[0].lower()),
        "name" : df["name"],
        #"value" : df["value"],
        #"frequency" : df["frequency"].apply(lambda x: str("0") if x is np.nan else frequecy_mapping[x.lower()]),
        "logo" : df["logo"],
        "homepage" : df["homepage"]
    })

    df_list = []
    for i, rows in df.iterrows():
        df_list.append({
            "name":rows['name'],
            "category":rows['type'],
            "superCategory":rows['type'],
            "logoUrl":rows['logo'] if rows['logo'] is not np.nan else "",
            "homePageUrl":rows['homepage'] if rows['homepage'] is not np.nan else "",
            "region":city
        })

    try:
        os.makedirs(settings.OUTPUT_PATH)
    except:
        pass
    with open('{}counterparty_pretty.json'.format(settings.OUTPUT_PATH), 'w') as outfile:
        json.dump(df_list, outfile, default=lambda x: x.dict(), indent=4)

@cli.command(help="init")
@click.option('--bank_number', default=3, help='bank_number')
@click.option('--branch_number', default=4, help='branch_number')
@click.option('--atm_number', default=6, help='atm_number')
@click.option('--product_number', default=10, help='product_number')
@click.option('--country', default='MXN', help='country')
@click.option('--output_path', default='./output_path', help='output_path')
def init(bank_number, branch_number, atm_number, product_number, country, output_dir):
    settings.BANK_NUMBER = str(bank_number)
    settings.BRANCH_NUMBER = str(branch_number)
    settings.ATM_NUMBER = str(atm_number)
    settings.PRODUCT_NUMBER = str(product_number)
    settings.COUNTRY=country
    settings.OUTPUT_PATH=output_dir

@cli.command(help="web_init")
@click.option('--api_host', default='http://127.0.0.1:8080', help='api_host')
@click.option('--redirect_url', default='http://127.0.0.1:9090', help='redirect_url')
@click.option('--admin_username', default='pflee', help='admin_username')
@click.option('--admin_password', default='Pflee@0218', help='admin_password')
@click.option('--file_root', default='G:/OBP-Project/SandboxDataGenerator/output_path', help='file_root')
def web_init(api_host, redirect_url, admin_username, admin_password, file_root):
    os.environ['API_HOST'] = str(api_host)
    os.environ['REDIRECT_URL'] = str(redirect_url)
    os.environ['ADMIN_USERNAME'] = str(admin_username)
    os.environ['ADMIN_PASSWORD'] = str(admin_password)
    os.environ['FILE_ROOT']=file_root
    click.echo('set successfully')
    click.echo('API_HOST: '.format(settings.API_HOST))
    click.echo('REDIRECT_URL: '.format(settings.REDIRECT_URL))
    click.echo('ADMIN_USERNAME: '.format(settings.ADMIN_USERNAME))
    click.echo('ADMIN_PASSWORD: '.format(settings.ADMIN_PASSWORD))
    click.echo('FILE_ROOT: '.format(settings.FILE_ROOT))

@cli.command(help="import_main")
def import_main():
    adminUserUsername = settings.ADMIN_USERNAME
    adminPassword = settings.ADMIN_PASSWORD
    from object.Admin import Admin
    admin_user = Admin(adminUserUsername, adminPassword)
    session = admin_user.direct_login()

    file_json = Admin.load(settings.FILE_ROOT+"sandbox_pretty.json")
    url = settings.API_HOST+"/obp/v3.0.0/sandbox/data-import"
    result2 = session.request('POST', url, json=file_json, verify=settings.VERIFY)
    if result2.status_code==201:
        click.echo("Import Successfully.")
    else:
        click.echo("Import Failed.")

@cli.command(help="import_customer")
def import_customer():
    adminUserUsername = settings.ADMIN_USERNAME
    adminPassword = settings.ADMIN_PASSWORD

    from object.PostCustomer import PostCustomer
    json_customers = PostCustomer.load(settings.FILE_ROOT+"/customers_pretty.json")

    click.echo("Got {} records".format(len(json_customers)))

    customer_list = [PostCustomer(customer['customer_number'],
                                  customer['legal_name'],
                                  customer['mobile_phone_number'],
                                  customer['email'],
                                  customer['face_image'],
                                  customer['date_of_birth'],
                                  customer['relationship_status'],
                                  customer['dependants'],
                                  customer['dob_of_dependants'],
                                  customer['highest_education_attained'],
                                  customer['employment_status'],
                                  customer['kyc_status'],
                                  customer['last_ok_date'],
                                  customer['bank_id'],
                                  customer['credit_rating'],
                                  customer['credit_limit']
                                  ) for customer in json_customers]

    from object.Admin import Admin
    from object.Bank_Import import Bank_Import
    json_user = Admin.load(settings.FILE_ROOT+"sandbox_pretty.json")
    click.echo("Got {} users".format(len(json_user['users'])))

    click.echo("login as user: ")
    admin_user = Admin(adminUserUsername, adminPassword)
    session = admin_user.direct_login()
    click.echo("login successfully!!!")

    click.echo("Got {} banks".format(len(json_user['banks'])))
    bank_list=[]
    for bank in json_user['banks']:
        bank_list.append(Bank_Import(bank['id'],
                              bank['short_name'],
                              bank['full_name'],
                              bank['logo'],
                              bank['website']))

    for user_dict in json_user['users']:
        user = Admin(user_dict['user_name'], user_dict['password'], user_dict['email'])
        customer_filtered = [customer for customer in customer_list if customer.email == user.email]
        result = session.get(
            settings.API_HOST + "/obp/v3.1.0/users/username/" + user.user_name)
        if result.status_code==200:
            current_user = json.loads(result.content)

            for customer in customer_filtered:
                click.echo("email is {} customer number is {} name is {} and has {} dependants born on {} "
                      .format(customer.email, customer.customer_number, customer.legal_name, customer.dependants, customer.dob_of_dependants))

                for bank in bank_list:
                    click.echo("Posting a customer for bank {}".format(bank.short_name))
                    url = settings.API_HOST+"/obp/v2.1.0/banks/{}/customers".format(bank.id)
                    result2 = session.request('POST', url, json=customer.to_json(current_user['user_id']), verify=settings.VERIFY)
                    if result2.status_code==201:
                        click.echo("saved {} as customer {}".format(customer.customer_number, result2.content))
                    else:
                        click.echo("did NOT save customer {}".format(result2.content if result2 is not None and result2.content is not None else ""))
        else:
            click.echo(result.content if result is not None and result.content is not None else "")

@cli.command(help="import_counterparty")
@click.option('--file_root', default='./output_path', help='file_root')
def import_counterparty(file_root):
    from object.Admin import Admin
    from object.PostCounterparty import PostCounterparty
    json_object_counterparty= PostCounterparty.load(settings.FILE_ROOT+"OBP_sandbox_counterparties_pretty_2.json")

    counterparty_list = [val for sublist in json_object_counterparty for val in sublist]

    json_object_user=Admin.load(settings.FILE_ROOT+"OBP_sandbox_pretty_2.json")

    for user_dict in json_object_user['users']:
        user = Admin(user_dict['user_name'], user_dict['password'], user_dict['email'])
        click.echo("login as user: ")
        session = user.direct_login()
        click.echo("get users private accounts")
        private_account = user.get_user_private_account()
        account_list = json.loads(private_account)['accounts']
        click.echo("ok!!!")

        click.echo("get other accounts for the accounts")
        for account in account_list:
            bank_id = account['bank_id']
            region = bank_id.split('.')[2]
            account_id = account['id']
            view = account['views_available'][0]
            result = user.get_user_other_account(bank_id, account_id, view['id'])
            click.echo(type(result))
            other_accounts_list = json.loads(result)['other_accounts']

            click.echo("bank_id: {}".format(bank_id))

            click.echo("region is {}".format(region))
            click.echo("get matching json counterparty data for each transaction's other_account")
            for other_accounts in other_accounts_list:
                counterparty_name = other_accounts['holder']['name']
                click.echo("Filtering counterparties by region {} and counterparty name {}".format(region, counterparty_name))
                regionCounterparties = [counterparty for counterparty in counterparty_list if counterparty['region']==region]
                records = [counterparty for counterparty in counterparty_list if counterparty['name'].lower()==counterparty_name.lower()]
                click.echo("Found {} records".format(len(records)))
                for cp in records:
                    click.echo("couterparty is Region {} Name {} Home Page {}".format(cp['region'],cp['name'],cp['homePageUrl']))
                    logoUrl = cp['homePageUrl'] if ("http://www.brandprofiles.com" in cp['logoUrl']) else cp['logoUrl']
                    if logoUrl.startswith("http") and other_accounts['metadata']['image_URL'] is None:
                        json_tmp = {"image_URL": logoUrl}
                        url = settings.API_HOST + "/obp/v3.1.0/banks/" + bank_id + "/accounts/" + account_id + "/" + view['id'] + "/other_accounts/"+other_accounts['id']+"/metadata/image_url"
                        result = session.request('POST', url, json = json_tmp, verify=settings.VERIFY)
                        if result.status_code == 201:
                            click.echo("Saved " + logoUrl + " as imageURL for counterparty "+ other_accounts['id'])
                        else:
                            click.echo("Save failed. {}".format(result.error if result is not None and result.error is not None else ""))
                    else:
                        click.echo("did NOT save " + logoUrl + " as imageURL for counterparty "+ other_accounts['id'])

                    if (cp['homePageUrl'].startswith("http") and not cp['homePageUrl'].endswith("jpg") and not cp['homePageUrl'].endswith("png") and other_accounts['metadata']['URL'] is None):
                        json_tmp = {"URL": cp['homePageUrl']}
                        url = settings.API_HOST + "/obp/v3.1.0/banks/" + bank_id + "/accounts/" + account_id + "/" + \
                              view['id'] + "/other_accounts/" + other_accounts['id'] + "/metadata/url"
                        result = session.request('POST', url, json=json_tmp, verify=settings.VERIFY)
                        if result.status_code == 201:
                            click.echo("Saved " + cp['homePageUrl'] + " as URL for counterparty "+ other_accounts['id'])
                        else:
                            click.echo("Save failed. {}".format(result.error if result is not None and result.error is not None else ""))
                    else:
                        click.echo("did NOT save " + cp['homePageUrl'] + " as URL for counterparty "+ other_accounts['id'])

                    if (cp['category'] is not None and other_accounts['metadata']['more_info'] is None):
                        categoryBits = cp['category'].split("_")
                        moreInfo = categoryBits[0]

                        json_tmp = {"more_info": moreInfo}
                        url = settings.API_HOST + "/obp/v3.1.0/banks/" + bank_id + "/accounts/" + account_id + "/" + view['id'] + "/other_accounts/" + other_accounts['id'] + "/metadata/more_info"
                        result = session.request('POST', url, json=json_tmp, verify=settings.VERIFY)
                        if result.status_code==201:
                            click.echo("Saved " + moreInfo + " as more_info for counterparty "+ other_accounts['id'])
                        else:
                            click.echo("Save failed. {}".format(result.error if result is not None and result.error is not None else ""))
                    else:
                        if other_accounts['metadata']['more_info'] is not None:
                            click.echo("more info is not empty:{}")
                        else:
                            click.echo("did NOT save more_info for counterparty "+ other_accounts['id'])

        user.oauth_logout()