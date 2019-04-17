from openpyxl import load_workbook
import re
import json
import settings
import numpy as np

if __name__=='__main__':
    file_path = "../input_file/OBP_counterparties_TESOBE_Hong_Kong.xlsx"
    wb = load_workbook(file_path)

    city = "Hong Kong"

    sheet_selected = [sheetname for sheetname in wb.sheetnames if city in sheetname ]

    import pandas as pd

    dataframe = pd.read_excel(file_path, sheet_name = sheet_selected, header=None, index_col=None, skiprows=3)

    df = pd.DataFrame()
    for i, j in dataframe.items():
        df = pd.DataFrame(j)

    df.columns= ['type', 'name', 'reference', 'value', 'frequency', 'logo','homepage']

    df.type = df["type"].apply(lambda x: re.split('_|/',x)[0].lower())
    #df.frequency = df["frequency"].apply(lambda x: str("0") if x is np.nan else frequecy_mapping[x.lower()])

    df = pd.DataFrame({
        "type":df["type"].apply(lambda x: re.split('_|/',x)[0].lower()),
        "name" : df["name"],
        #"value" : df["value"],
        #"frequency" : df["frequency"].apply(lambda x: str("0") if x is np.nan else frequecy_mapping[x.lower()]),
        "logo" : df["logo"],
        "homepage" : df["homepage"]
    })

    df_list = []
    for i, rows in df.iterrows():
        df_list.append({
            "name":rows['name'],
            "category":rows['type'],
            "superCategory":rows['type'],
            "logoUrl":rows['logo'] if rows['logo'] is not np.nan else "",
            "homePageUrl":rows['homepage'] if rows['homepage'] is not np.nan else "",
            "region":city
        })

    with open('{}counterparty_pretty.json'.format(settings.OUTPUT_PATH), 'w') as outfile:
        json.dump(df_list, outfile, default=lambda x: x.dict(), indent=4)