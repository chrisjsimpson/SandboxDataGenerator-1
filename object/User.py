import random
import settings
import pandas as pd
from openpyxl import load_workbook

from object.Account import Account
from object.Customer import Customer
from object.Routing import *

boy_first_names=('John','Andy','Joe')
girl_first_names=('Alice','Ammy')
last_names=('Johnson','Smith','Williams')
mails=('gmail.com','qq.com','tesobe.com')
class User:
    def __init__(self, username, email, password, phone, relationship_status,
                 employment_status, highest_education_attained,
                 savings, current, country, food, utility, clothing,auto,
            health,	entertainment,	gift,	education,	fee,	rent

    ):
        self.username = username
        self.email = email
        self.password = password
        self.phone = phone
        self.relationship_status = relationship_status
        self.employment_status = employment_status
        self.highest_education_attained = highest_education_attained
        self.savings = savings
        self.current = current
        self.country = country
        self.food = food
        self.utility = utility
        self.clothing = clothing
        self.auto = auto
        self.health = health
        self.entertainment = entertainment
        self.gift = gift
        self.education = education
        self.fee = fee
        self.rent = rent

    def dict(self):
        return {
            'user_name':self.username,
            'email': self.email,
            'password': self.password
        }

    def create_account(self, branch, type, currency, amount):
        account_id=str(uuid.uuid4())
        number = str(digits11())
        label = "{} {}".format(self.username, branch.name)
        iban = generateIBAN(branch.address.country_code.upper())
        return Account(
            account_id=str(uuid.uuid4()),
            branch=branch,
            user=[self],
            label=label,
            number=number,
            type=type,
            balance={
                 "currency":currency,
                 "amount":amount
             },
            account_routing=Routing(
                scheme="OBP",
                address=account_id
            ),
            iban=iban
        )

    def create_customer(self, bank):
        customer_number = str(uuid.uuid4())
        face_image = {
            "url": "www.example.com",
            "date": gen_datetime().strftime("%Y-%m-%dT%H:%M:%SZ")
        }
        date_of_birth = gen_datetime(min_year = 1970, max_year=2001).strftime("%Y-%m-%dT%H:%M:%SZ")
        dob_of_dependants = [
            gen_datetime().strftime("%Y-%m-%dT%H:%M:%SZ") for _ in range(2)
        ]

        kyc_status = True
        last_ok_date = gen_datetime().strftime("%Y-%m-%dT%H:%M:%SZ")
        credit_rating=""
        credit_limit=""
        return Customer(
            customer_number=customer_number,
            user=self,
            mobile_phone_number=self.phone,
            face_image=face_image,
            date_of_birth=date_of_birth,
            relationship_status=self.relationship_status,
            dependants=random.randint(0,6),
            dob_of_dependants=dob_of_dependants,
            highest_education_attained=self.highest_education_attained,
            employment_status=self.employment_status,
            kyc_status=kyc_status,
            last_ok_date=last_ok_date,
            bank=bank,
            credit_rating=credit_rating,
            credit_limit=credit_limit
        )

    @staticmethod
    def Generator(num, gender = 'male', relationship_status = 'married',
                  employment_status = 'retired', highest_education_attained='BA.',
                  savings = 80000, current = 2400, country = 'MXN',
                  food=5, utility=0.2, clothing=1,auto=0.2,
                  health=0.1,	entertainment=1,	gift=0.2,	education=0.2,	fee=0.5,	rent=0,
                  input_file = settings.DATASET_PATH):
        wb = load_workbook(input_file)

        if gender=='male':
            first_names = boy_first_names
        else:
            first_names = girl_first_names
        first_name_tmp = random.choice(first_names)
        last_name = random.choice(last_names)

        if 'first_name' in wb.sheetnames:
            dataframe = pd.read_excel(input_file, sheet_name = 'first_name', header=0, index_col=None)
            if gender == 'male':
                first_names = dataframe.Boy.values
            else:
                first_names = dataframe.Girl.values
            first_name_tmp = random.choice(first_names)

        if 'last_name' in wb.sheetnames:
            dataframe = pd.read_excel(input_file, sheet_name = 'last_name', header=None, index_col=None)
            last_name = random.choice(dataframe[0].values)

        for i in range(num):
            first_name = "{}".format(first_name_tmp)
            username = "{} {}".format(first_name, last_name)
            email = first_name +"@"+ random.choice(mails)
            password = str(uuid.uuid4())[:10]
            yield User(
                username=username,
                email=email,
                password=password,
                phone=phone_number_generation(),
                relationship_status= relationship_status,
                employment_status= employment_status,
                highest_education_attained = highest_education_attained,
                savings=savings,
                current=current,
                country=country,
                food = food,
                utility = utility,
                clothing = clothing,
                auto = auto,
                health = health,
                entertainment = entertainment,
                gift = gift,
                education = education,
                fee = fee,
                rent = rent
            )

    @staticmethod
    def generator_for_file(input_file = settings.OPTIONS_PATH):
        wb = load_workbook(input_file)

        if 'user' in wb.sheetnames:
            dataframe = pd.read_excel(input_file, sheet_name = 'user', header=0, index_col=None)
            users = []
            for row in dataframe.iterrows():
                row = row[1]
                users.append(next(User.Generator(1, row['gender'], row['relationship_status'], row['employment_status'],
                                                 row['highest_education_attained'], row['savings'],
                                                 row['current'], row['country'],
                                                 row['food'], row['utility'],
                                                 row['clothing'], row['auto'],
                                                 row['health'], row['entertainment'],
                                                 row['gift'], row['education'],
                                                 row['fee'], row['rent'],
                                                 )))
            return users
        else:
            raise Exception("There is not user option.")